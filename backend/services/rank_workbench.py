from __future__ import annotations

import base64
import csv
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.db import DATA_DIR
from backend.services.rank_core import RankResult, choose_provider, estimate_credit_cost

RANK_EXPORT_DIR = DATA_DIR / "rank_exports"
RANK_EXPORT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class TemplateWorksheet:
    sheet_name: str
    keyword_rows: list[tuple[int, str]]
    history_headers: list[str]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decode_file_content(content_base64: str) -> bytes:
    return base64.b64decode(content_base64.encode("utf-8"))


def _load_workbook_bytes(content_base64: str):
    payload = _decode_file_content(content_base64)
    return load_workbook(filename=BytesIO(payload))


def _pick_template_sheet(workbook) -> TemplateWorksheet:
    for worksheet in workbook.worksheets:
        keyword_rows: list[tuple[int, str]] = []
        for row_idx in range(2, worksheet.max_row + 1):
            keyword = _clean_text(worksheet.cell(row=row_idx, column=1).value)
            if keyword:
                keyword_rows.append((row_idx, keyword))
        if keyword_rows:
            history_headers = [
                _clean_text(worksheet.cell(row=1, column=column_idx).value)
                for column_idx in range(2, worksheet.max_column + 1)
                if _clean_text(worksheet.cell(row=1, column=column_idx).value)
            ]
            return TemplateWorksheet(
                sheet_name=worksheet.title,
                keyword_rows=keyword_rows,
                history_headers=history_headers,
            )
    worksheet = workbook[workbook.sheetnames[0]]
    return TemplateWorksheet(sheet_name=worksheet.title, keyword_rows=[], history_headers=[])


def preview_rank_template(filename: str, content_base64: str) -> dict[str, Any]:
    workbook = _load_workbook_bytes(content_base64)
    template_sheet = _pick_template_sheet(workbook)
    return {
        "filename": filename,
        "sheet_name": template_sheet.sheet_name,
        "keyword_count": len(template_sheet.keyword_rows),
        "history_column_count": len(template_sheet.history_headers),
        "history_headers": template_sheet.history_headers,
        "keyword_preview": [keyword for _, keyword in template_sheet.keyword_rows[:10]],
    }


def _display_rank(result: dict[str, Any]) -> str:
    if result.get("found") and result.get("page") and result.get("position"):
        return f"{result['page']}--{result['position']}"
    return "/--/"


def _timestamp_slug() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _date_column_label(existing_headers: list[str]) -> str:
    base = datetime.now().strftime("%Y-%m-%d")
    if base not in existing_headers:
        return base
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _build_result_payload(result: RankResult) -> dict[str, Any]:
    return {
        "keyword": result.keyword,
        "found": result.found,
        "page": result.page,
        "position": result.position,
        "display_rank": f"{result.page}--{result.position}" if result.found and result.page and result.position else "/--/",
        "url": result.url or "",
        "provider": result.provider,
        "error": result.error or "",
        "queried_at": result.queried_at or "",
    }


def _matrix_rows_from_worksheet(worksheet, latest_result_map: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    columns = [_clean_text(worksheet.cell(row=1, column=column_idx).value) for column_idx in range(2, worksheet.max_column + 1)]
    columns = [column for column in columns if column]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, worksheet.max_row + 1):
        keyword = _clean_text(worksheet.cell(row=row_idx, column=1).value)
        if not keyword:
            continue
        values: list[str] = []
        for column_idx in range(2, worksheet.max_column + 1):
            value = _clean_text(worksheet.cell(row=row_idx, column=column_idx).value)
            if value == "" and column_idx - 2 < len(columns):
                value = "/--/"
            if column_idx - 2 < len(columns):
                values.append(value)
        latest_value = values[-1] if values else ""
        misses_streak = 0
        for item in reversed(values):
            if item and item != "/--/":
                break
            misses_streak += 1
        rows.append(
            {
                "keyword": keyword,
                "values": values,
                "latest_value": latest_value,
                "latest_found": bool(latest_value and latest_value != "/--/"),
                "miss_streak": misses_streak,
                "history_items": [{"column": column, "value": values[index]} for index, column in enumerate(columns) if index < len(values)],
                "latest_result": latest_result_map.get(keyword),
            }
        )
    return rows


def _save_detail_csv(results: list[dict[str, Any]], path: Path) -> None:
    headers = ["keyword", "found", "page", "position", "display_rank", "url", "provider", "error", "queried_at"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for item in results:
            writer.writerow({key: item.get(key, "") for key in headers})


def run_batch_template_tracking(
    *,
    filename: str,
    content_base64: str,
    domain: str,
    provider_name: str,
    max_pages: int,
    results_per_request: int,
    hl: str,
    gl: str,
    api_key: str,
) -> dict[str, Any]:
    workbook = _load_workbook_bytes(content_base64)
    template_sheet = _pick_template_sheet(workbook)
    worksheet = workbook[template_sheet.sheet_name]

    if not template_sheet.keyword_rows:
        raise RuntimeError("The uploaded template does not contain keywords in the first column.")

    provider = choose_provider(
        provider_name=provider_name,
        api_key=api_key,
        hl=hl,
        gl=gl,
        results_per_request=results_per_request,
    )
    account_info = provider.get_account_info()

    keywords = [keyword for _, keyword in template_sheet.keyword_rows]
    if account_info and account_info.total_searches_left is not None:
        per_keyword_cost = max(1, estimate_credit_cost(1, max_pages, results_per_request, provider_name))
        allowed_keywords = max(0, account_info.total_searches_left // per_keyword_cost)
        keywords = keywords[:allowed_keywords]
        template_sheet.keyword_rows = template_sheet.keyword_rows[: len(keywords)]

    date_header = _date_column_label(template_sheet.history_headers)
    date_column = worksheet.max_column + 1
    worksheet.cell(row=1, column=date_column, value=date_header)

    results: list[dict[str, Any]] = []
    latest_result_map: dict[str, dict[str, Any]] = {}
    for row_idx, keyword in template_sheet.keyword_rows:
        queried_at = datetime.now().astimezone().isoformat(timespec="seconds")
        try:
            result = provider.find_rank(keyword=keyword, target_domain=domain, max_pages=max_pages)
            result.queried_at = queried_at
        except Exception as exc:
            result = RankResult(
                keyword=keyword,
                found=False,
                provider=provider.name,
                error=str(exc),
                queried_at=queried_at,
            )
        item = _build_result_payload(result)
        worksheet.cell(row=row_idx, column=date_column, value=item["display_rank"])
        results.append(item)
        latest_result_map[keyword] = item

    input_name = Path(filename).stem or "rank-template"
    slug = _timestamp_slug()
    output_xlsx = RANK_EXPORT_DIR / f"{input_name}_rank_history_{slug}.xlsx"
    output_csv = RANK_EXPORT_DIR / f"{input_name}_rank_details_{slug}.csv"
    workbook.save(output_xlsx)
    _save_detail_csv(results, output_csv)

    columns = template_sheet.history_headers + [date_header]
    rows = _matrix_rows_from_worksheet(worksheet, latest_result_map)
    summary = {
        "total": len(results),
        "found": sum(1 for item in results if item["found"]),
        "errors": sum(1 for item in results if item["error"]),
        "notFound": sum(1 for item in results if not item["found"]),
        "mode": "batch_template_run",
        "input_file": filename,
        "output_file": str(output_xlsx),
        "detail_file": str(output_csv),
        "sheet_name": template_sheet.sheet_name,
        "new_date_column": date_header,
        "keyword_count": len(template_sheet.keyword_rows),
        "template_preview": {
            "filename": filename,
            "sheet_name": template_sheet.sheet_name,
            "keyword_count": len(template_sheet.keyword_rows),
            "history_column_count": len(template_sheet.history_headers),
            "history_headers": template_sheet.history_headers,
            "keyword_preview": keywords[:10],
        },
        "matrix": {
            "columns": columns,
            "rows": rows,
        },
    }
    return {
        "mode": "batch_template_run",
        "results": results,
        "summary": summary,
        "output_file": str(output_xlsx),
        "detail_file": str(output_csv),
        "columns": columns,
        "rows": rows,
        "template_preview": summary["template_preview"],
    }


def run_single_keyword_tracking(
    *,
    keyword: str,
    domain: str,
    provider_name: str,
    max_pages: int,
    results_per_request: int,
    hl: str,
    gl: str,
    api_key: str,
) -> dict[str, Any]:
    provider = choose_provider(
        provider_name=provider_name,
        api_key=api_key,
        hl=hl,
        gl=gl,
        results_per_request=results_per_request,
    )
    queried_at = datetime.now().astimezone().isoformat(timespec="seconds")
    try:
        result = provider.find_rank(keyword=keyword, target_domain=domain, max_pages=max_pages)
        result.queried_at = queried_at
    except Exception as exc:
        result = RankResult(
            keyword=keyword,
            found=False,
            provider=provider.name,
            error=str(exc),
            queried_at=queried_at,
        )
    item = _build_result_payload(result)
    return {
        "mode": "single_keyword_check",
        "results": [item],
        "summary": {
            "total": 1,
            "found": 1 if item["found"] else 0,
            "errors": 1 if item["error"] else 0,
            "notFound": 0 if item["found"] else 1,
            "mode": "single_keyword_check",
        },
    }
