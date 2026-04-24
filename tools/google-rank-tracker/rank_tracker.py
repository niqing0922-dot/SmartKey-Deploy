import argparse
import csv
import os
import random
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from urllib.parse import parse_qs, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = str(BASE_DIR / "IoT Keywords - 1.xlsx")
DEFAULT_SHEET = "Keywords_Clean"
DEFAULT_DOMAIN = "waveteliot.com"
DEFAULT_OUTPUT = "waveteliot_google_rankings.xlsx"
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
]


@dataclass
class RankResult:
    keyword: str
    found: bool
    page: Optional[int] = None
    position: Optional[int] = None
    url: Optional[str] = None
    provider: str = ""
    error: Optional[str] = None
    queried_at: Optional[str] = None


@dataclass
class AccountInfo:
    total_searches_left: Optional[int] = None
    plan_searches_left: Optional[int] = None
    searches_per_month: Optional[int] = None
    this_month_usage: Optional[int] = None


def domain_matches(url: str, target_domain: str) -> bool:
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return False
    host = host.split(":")[0]
    target = target_domain.lower()
    return host == target or host.endswith(f".{target}")


def extract_keywords(
    xlsx_path: Path,
    sheet_name: Optional[str] = None,
    column: int = 1,
    skip_rows: int = 1,
) -> List[str]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    keywords: List[str] = []
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_col=column, max_col=column, values_only=True),
        start=1,
    ):
        if row_idx <= skip_rows:
            continue
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        keywords.append(text)
    return keywords


class SearchProvider:
    name = "base"

    def find_rank(self, keyword: str, target_domain: str, max_pages: int) -> RankResult:
        raise NotImplementedError

    def get_account_info(self) -> Optional[AccountInfo]:
        return None


class SerpApiProvider(SearchProvider):
    name = "serpapi"

    def __init__(self, api_key: str, hl: str = "en", gl: str = "us", results_per_request: int = 100) -> None:
        self.api_key = api_key
        self.hl = hl
        self.gl = gl
        self.results_per_request = max(10, min(results_per_request, 100))
        self.session = requests.Session()

    def find_rank(self, keyword: str, target_domain: str, max_pages: int) -> RankResult:
        search_results_per_page = 10
        request_count = max(1, (max_pages * search_results_per_page + self.results_per_request - 1) // self.results_per_request)
        for request_index in range(request_count):
            start = request_index * self.results_per_request
            params = {
                "engine": "google",
                "q": keyword,
                "num": self.results_per_request,
                "start": start,
                "api_key": self.api_key,
            }
            if self.hl:
                params["hl"] = self.hl
            if self.gl:
                params["gl"] = self.gl
            response = self.session.get(
                "https://serpapi.com/search.json",
                params=params,
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("error"):
                raise RuntimeError(payload["error"])

            results = payload.get("organic_results", [])
            for idx, item in enumerate(results, start=1):
                link = item.get("link") or item.get("redirect_link")
                absolute_position = start + idx
                if link and domain_matches(link, target_domain):
                    return RankResult(
                        keyword=keyword,
                        found=True,
                        page=((absolute_position - 1) // search_results_per_page) + 1,
                        position=absolute_position,
                        url=link,
                        provider=self.name,
                    )

            if not results:
                break

        return RankResult(keyword=keyword, found=False, provider=self.name)

    def get_account_info(self) -> Optional[AccountInfo]:
        response = self.session.get(
            "https://serpapi.com/account.json",
            params={"api_key": self.api_key},
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        return AccountInfo(
            total_searches_left=payload.get("total_searches_left"),
            plan_searches_left=payload.get("plan_searches_left"),
            searches_per_month=payload.get("searches_per_month"),
            this_month_usage=payload.get("this_month_usage"),
        )


class GoogleHtmlProvider(SearchProvider):
    name = "google_html"

    def __init__(self, hl: str = "en", gl: str = "us", delay_range: tuple[float, float] = (2.0, 4.0)) -> None:
        self.hl = hl
        self.gl = gl
        self.delay_range = delay_range
        self.session = requests.Session()

    def _headers(self) -> dict:
        return {"User-Agent": random.choice(USER_AGENTS)}

    def _extract_links(self, html: str) -> Iterable[str]:
        soup = BeautifulSoup(html, "html.parser")
        seen = set()
        for anchor in soup.select("a[href]"):
            href = anchor.get("href", "")
            if href.startswith("/url?"):
                parsed = parse_qs(urlparse(href).query)
                url = parsed.get("q", [None])[0]
                if url and url not in seen:
                    seen.add(url)
                    yield url

    def find_rank(self, keyword: str, target_domain: str, max_pages: int) -> RankResult:
        per_page = 10
        for page in range(1, max_pages + 1):
            start = (page - 1) * per_page
            params = {
                "q": keyword,
                "num": per_page,
                "start": start,
                "pws": "0",
            }
            if self.hl:
                params["hl"] = self.hl
            if self.gl:
                params["gl"] = self.gl
            response = self.session.get(
                "https://www.google.com/search",
                params=params,
                headers=self._headers(),
                timeout=30,
            )
            response.raise_for_status()
            html = response.text

            if "enablejs" in html or "unusual traffic" in html.lower():
                raise RuntimeError("Google 返回了反爬/启用 JavaScript 页面，建议改用 SerpAPI 模式。")

            links = list(self._extract_links(html))
            for idx, link in enumerate(links[:per_page], start=1):
                absolute_position = start + idx
                if domain_matches(link, target_domain):
                    return RankResult(
                        keyword=keyword,
                        found=True,
                        page=page,
                        position=absolute_position,
                        url=link,
                        provider=self.name,
                    )

            time.sleep(random.uniform(*self.delay_range))

        return RankResult(keyword=keyword, found=False, provider=self.name)


def choose_provider(provider_name: str, hl: str, gl: str, results_per_request: int) -> SearchProvider:
    provider_name = provider_name.lower()
    if provider_name == "serpapi":
        api_key = os.getenv("SERPAPI_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError("未设置环境变量 SERPAPI_API_KEY。")
        return SerpApiProvider(api_key=api_key, hl=hl, gl=gl, results_per_request=results_per_request)
    if provider_name == "google_html":
        return GoogleHtmlProvider(hl=hl, gl=gl)
    raise ValueError(f"不支持的 provider: {provider_name}")


def save_results(results: List[RankResult], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "rankings"
    headers = ["keyword", "found", "page", "position", "url", "provider", "error", "queried_at"]
    sheet.append(headers)
    for item in results:
        sheet.append(
            [
                item.keyword,
                "yes" if item.found else "no",
                item.page,
                item.position,
                item.url,
                item.provider,
                item.error,
                item.queried_at,
            ]
        )
    workbook.save(output_path)

    csv_path = output_path.with_suffix(".csv")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for item in results:
            writer.writerow(
                [
                    item.keyword,
                    "yes" if item.found else "no",
                    item.page,
                    item.position,
                    item.url,
                    item.provider,
                    item.error,
                    item.queried_at,
                ]
            )


def resolve_output_path(input_path: Path, output_arg: Optional[str]) -> Path:
    if output_arg:
        candidate = Path(output_arg)
        if candidate.is_absolute():
            return candidate
        return input_path.parent / candidate
    dated_output = datetime.now().strftime("waveteliot_google_rankings_%Y-%m-%d.xlsx")
    return input_path.parent / dated_output


def load_existing_results(output_path: Path) -> Dict[str, RankResult]:
    if not output_path.exists():
        return {}

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    existing: Dict[str, RankResult] = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        keyword = str(row[0]).strip()
        if not keyword:
            continue
        existing[keyword] = RankResult(
            keyword=keyword,
            found=str(row[1]).strip().lower() == "yes" if row[1] is not None else False,
            page=int(row[2]) if row[2] is not None else None,
            position=int(row[3]) if row[3] is not None else None,
            url=str(row[4]).strip() if row[4] else None,
            provider=str(row[5]).strip() if row[5] else "",
            error=str(row[6]).strip() if row[6] else None,
            queried_at=str(row[7]).strip() if len(row) > 7 and row[7] else None,
        )
    return existing


def can_resume_result(result: RankResult, provider_name: str) -> bool:
    if result.error:
        return False
    return result.provider == provider_name


def estimate_credit_cost(keyword_count: int, max_pages: int, results_per_request: int, provider_name: str) -> int:
    if provider_name != "serpapi":
        return 0
    search_results_per_page = 10
    requests_per_keyword = max(1, (max_pages * search_results_per_page + results_per_request - 1) // results_per_request)
    return keyword_count * requests_per_keyword


def main() -> None:
    parser = argparse.ArgumentParser(description="查询关键词在 Google 中的目标域名排名。")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="关键词 Excel 路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="工作表名称，默认 Keywords_Clean")
    parser.add_argument("--column", type=int, default=1, help="关键词所在列，默认第 1 列")
    parser.add_argument("--skip-rows", type=int, default=1, help="从顶部跳过多少行，默认 1")
    parser.add_argument("--domain", default=DEFAULT_DOMAIN, help="目标域名")
    parser.add_argument("--provider", default="serpapi", choices=["serpapi", "google_html"], help="搜索结果来源")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 个关键词，0 表示全部")
    parser.add_argument("--max-pages", type=int, default=10, help="最多检查多少页，默认 10")
    parser.add_argument("--results-per-request", type=int, default=100, help="SerpAPI 每次请求拉取多少条结果，默认 100")
    parser.add_argument("--resume", action="store_true", help="如果结果文件已存在，则跳过已完成的关键词")
    parser.add_argument("--reserve-credits", type=int, default=10, help="SerpAPI 至少保留多少剩余额度，默认 10")
    parser.add_argument("--hl", default="en", help="Google 语言参数，例如 en；留空则不传")
    parser.add_argument("--gl", default="", help="Google 国家参数，例如 us；留空则不传")
    parser.add_argument("--output", default=None, help="输出文件名，默认写入 Excel 所在文件夹")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = resolve_output_path(input_path, args.output)

    keywords = extract_keywords(
        input_path,
        sheet_name=args.sheet,
        column=args.column,
        skip_rows=args.skip_rows,
    )
    if not keywords:
        raise RuntimeError("没有从 Excel 中读取到关键词。")
    if args.limit > 0:
        keywords = keywords[: args.limit]

    provider = choose_provider(args.provider, hl=args.hl, gl=args.gl, results_per_request=args.results_per_request)
    all_keywords = list(keywords)
    loaded_existing = load_existing_results(output_path) if args.resume else {}
    output_existing = {
        keyword: result
        for keyword, result in loaded_existing.items()
        if can_resume_result(result, args.provider)
    }
    if output_existing:
        keywords = [keyword for keyword in keywords if keyword not in output_existing]
    results_map: Dict[str, RankResult] = {
        keyword: result
        for keyword, result in loaded_existing.items()
        if keyword not in all_keywords or can_resume_result(result, args.provider)
    }

    if args.provider == "serpapi":
        account_info = provider.get_account_info()
        if account_info and account_info.total_searches_left is not None:
            estimated_cost = estimate_credit_cost(
                keyword_count=len(keywords),
                max_pages=args.max_pages,
                results_per_request=args.results_per_request,
                provider_name=args.provider,
            )
            allowed = max(0, account_info.total_searches_left - args.reserve_credits)
            print(
                f"SerpAPI credits left: {account_info.total_searches_left}, "
                f"reserved: {args.reserve_credits}, estimated cost: {estimated_cost}"
            )
            per_keyword_cost = max(1, estimate_credit_cost(1, args.max_pages, args.results_per_request, args.provider))
            max_keywords_by_budget = allowed // per_keyword_cost if per_keyword_cost else len(keywords)
            if max_keywords_by_budget < len(keywords):
                keywords = keywords[:max_keywords_by_budget]
                print(f"Budget limit applied, processing {len(keywords)} keyword(s) this run.")

    print(f"Loaded {len(keywords)} keywords from {input_path}")
    print(f"Provider: {provider.name}, target domain: {args.domain}, max pages: {args.max_pages}")

    for index, keyword in enumerate(keywords, start=1):
        print(f"[{index}/{len(keywords)}] Searching: {keyword}")
        queried_at = datetime.now().astimezone().isoformat(timespec="seconds")
        try:
            result = provider.find_rank(keyword=keyword, target_domain=args.domain, max_pages=args.max_pages)
        except Exception as exc:
            result = RankResult(
                keyword=keyword,
                found=False,
                provider=provider.name,
                error=str(exc),
                queried_at=queried_at,
            )
        else:
            result.queried_at = queried_at
        results_map[keyword] = result

        if result.found:
            print(f"  -> page {result.page}, position {result.position}, url={result.url}")
        else:
            reason = f", error={result.error}" if result.error else ""
            print(f"  -> not found within top {args.max_pages * 10}{reason}")

    results = [results_map[keyword] for keyword in all_keywords if keyword in results_map]
    save_results(results, output_path)
    print(f"Saved: {output_path.resolve()}")
    print(f"Saved: {output_path.with_suffix('.csv').resolve()}")


if __name__ == "__main__":
    main()
